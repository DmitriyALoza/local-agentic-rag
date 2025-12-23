"""
Excel (XLSX) Document Parser
Extracts data from sheets, tables, and headers
"""

from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook


class ExcelParser:
    """Parser for Excel (.xlsx) files"""

    def parse(self, file_path: str) -> List[Dict]:
        """
        Parse an Excel file and extract structured content

        Args:
            file_path: Path to the .xlsx file

        Returns:
            List of dictionaries, each containing:
            - text: The extracted text content (formatted table data)
            - section: Section identifier (sheet name and range)
            - metadata: Additional metadata about the content
        """
        chunks = []
        filename = Path(file_path).name

        # Load workbook to get sheet names
        wb = load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            # Read sheet with pandas for easier data handling
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Skip empty sheets
                if df.empty:
                    continue

                # Convert dataframe to text representation
                # Include column headers and data
                text_parts = []

                # Add headers
                headers = " | ".join(str(col) for col in df.columns)
                text_parts.append(f"Headers: {headers}")

                # Add rows (limit to prevent huge chunks)
                max_rows = 100  # Configurable limit
                for idx, row in df.head(max_rows).iterrows():
                    row_text = " | ".join(str(val) for val in row.values)
                    text_parts.append(row_text)

                sheet_text = "\n".join(text_parts)

                # Create chunk for this sheet
                chunks.append({
                    "text": sheet_text,
                    "section": f"Sheet: {sheet_name}",
                    "metadata": {
                        "filename": filename,
                        "sheet_name": sheet_name,
                        "rows": len(df),
                        "columns": len(df.columns),
                        "column_names": df.columns.tolist(),
                        "content_type": "table"
                    }
                })

                # If sheet is too large, create additional chunks for remaining rows
                if len(df) > max_rows:
                    remaining_rows = len(df) - max_rows
                    for chunk_start in range(max_rows, len(df), max_rows):
                        chunk_df = df.iloc[chunk_start:chunk_start + max_rows]
                        chunk_text_parts = [f"Headers: {headers}"]

                        for idx, row in chunk_df.iterrows():
                            row_text = " | ".join(str(val) for val in row.values)
                            chunk_text_parts.append(row_text)

                        chunk_text = "\n".join(chunk_text_parts)

                        chunks.append({
                            "text": chunk_text,
                            "section": f"Sheet: {sheet_name} (rows {chunk_start + 1}-{min(chunk_start + max_rows, len(df))})",
                            "metadata": {
                                "filename": filename,
                                "sheet_name": sheet_name,
                                "row_start": chunk_start + 1,
                                "row_end": min(chunk_start + max_rows, len(df)),
                                "columns": len(df.columns),
                                "column_names": df.columns.tolist(),
                                "content_type": "table"
                            }
                        })

            except Exception as e:
                # If sheet cannot be read, create an error chunk
                chunks.append({
                    "text": f"Error reading sheet: {str(e)}",
                    "section": f"Sheet: {sheet_name} (Error)",
                    "metadata": {
                        "filename": filename,
                        "sheet_name": sheet_name,
                        "error": str(e),
                        "content_type": "error"
                    }
                })

        wb.close()
        return chunks

    def get_metadata(self, file_path: str) -> Dict:
        """
        Get metadata about the Excel file

        Args:
            file_path: Path to the .xlsx file

        Returns:
            Dictionary containing file metadata
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        path = Path(file_path)

        sheet_info = []
        for sheet_name in wb.sheetnames:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_info.append({
                    "name": sheet_name,
                    "rows": len(df),
                    "columns": len(df.columns)
                })
            except:
                sheet_info.append({
                    "name": sheet_name,
                    "error": "Unable to read sheet"
                })

        wb.close()

        return {
            "filename": path.name,
            "file_type": "xlsx",
            "total_sheets": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheet_info": sheet_info,
            "file_size_bytes": path.stat().st_size
        }
